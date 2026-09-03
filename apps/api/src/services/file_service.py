import csv
import io
import json
import os
import re
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import pypdf
import structlog
from anyio import Path as AsyncPath
from models.workspace import Workspace, WorkspaceFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

logger = structlog.get_logger()

# Allowed image MIME types and max payload limits
ALLOWED_IMAGE_MIMES = {
    "image/png",
    "image/jpeg",
    "image/jpg",
    "image/webp",
    "image/gif",
    "image/svg+xml",
}

TABULAR_EXTENSIONS = {
    ".csv",
    ".tsv",
    ".jsonl",
    ".ndjson",
    ".xlsx",
}

CODE_EXTENSIONS = {
    ".py",
    ".js",
    ".ts",
    ".tsx",
    ".jsx",
    ".json",
    ".yaml",
    ".yml",
    ".toml",
    ".sql",
    ".html",
    ".css",
    ".scss",
    ".sass",
    ".sh",
    ".bash",
    ".zsh",
    ".rs",
    ".go",
    ".c",
    ".cpp",
    ".h",
    ".hpp",
    ".java",
    ".kt",
    ".rb",
    ".php",
    ".cs",
    ".swift",
    ".dockerfile",
    ".graphql",
    ".proto",
    ".vue",
    ".svelte",
}

TEXT_DOC_EXTENSIONS = {
    ".txt",
    ".md",
    ".markdown",
    ".log",
    ".env",
    ".ini",
    ".cfg",
    ".conf",
    ".xml",
}

MAX_FILE_SIZE_BYTES = 20 * 1024 * 1024  # 20 Megabytes

# Storage root: data/storage/workspaces
BASE_STORAGE_DIR = Path(os.getenv("STORAGE_DIR", "data/storage"))


def _sanitize_filename(filename: str) -> str:
    """Sanitize filename to avoid path traversal or unsafe characters."""
    clean = os.path.basename(filename).strip()
    clean = re.sub(r"[^\w\.\-\_]", "_", clean)
    return clean or "uploaded_asset"


def parse_tabular_bytes(
    filename: str, content: bytes, mime_type: str = ""
) -> tuple[Optional[str], Dict[str, Any]]:
    """
    Parse CSV, TSV, JSONL, or XLSX content into a structured Markdown preview table
    and metadata dictionary (total rows, columns, format, etc.).
    """
    ext = os.path.splitext(filename.lower())[1]
    norm_mime = mime_type.lower().split(";")[0].strip()

    # 1. Excel (.xlsx)
    if ext == ".xlsx" or "spreadsheetml" in norm_mime:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            if not sheet_names:
                return None, {}
            active_sheet = wb.active or wb[sheet_names[0]]

            raw_rows: List[List[str]] = []
            for row in active_sheet.iter_rows(values_only=True):
                if any(v is not None and str(v).strip() != "" for v in row):
                    raw_rows.append([str(v) if v is not None else "" for v in row])

            if not raw_rows:
                empty_msg = f"**Dataset:** `{filename}` | **Sheet:** `{active_sheet.title}`\n*(Empty spreadsheet)*"
                return empty_msg, {
                    "format": "xlsx",
                    "sheets": sheet_names,
                    "active_sheet": active_sheet.title,
                    "row_count": 0,
                    "column_count": 0,
                    "columns": [],
                }

            headers = [h.strip() or f"Column_{idx + 1}" for idx, h in enumerate(raw_rows[0])]
            data_rows = raw_rows[1:]
            total_rows = len(data_rows)
            preview_rows = data_rows[:25]

            md_lines = [
                f"**Dataset:** `{filename}` | **Sheet:** `{active_sheet.title}`",
                f"**Format:** Excel (XLSX) | **Sheets:** {', '.join(sheet_names)} | **Total Rows:** {total_rows:,} | **Columns ({len(headers)}):** {', '.join(headers[:15])}",
                "",
                "| " + " | ".join(headers) + " |",
                "| " + " | ".join(["---"] * len(headers)) + " |",
            ]
            for r in preview_rows:
                padded = (r + [""] * len(headers))[: len(headers)]
                cleaned = [cell.replace("|", "\\|").replace("\n", " ").strip() for cell in padded]
                md_lines.append("| " + " | ".join(cleaned) + " |")

            if total_rows > len(preview_rows):
                md_lines.append(
                    f"\n*(Showing top {len(preview_rows)} of {total_rows:,} total rows)*"
                )

            return "\n".join(md_lines), {
                "format": "xlsx",
                "sheets": sheet_names,
                "active_sheet": active_sheet.title,
                "row_count": total_rows,
                "column_count": len(headers),
                "columns": headers,
            }
        except Exception as e:
            logger.warning("Failed to parse XLSX file", filename=filename, error=str(e))
            return None, {"error": str(e)}

    # 2. JSONL (.jsonl, .ndjson)
    if ext in (".jsonl", ".ndjson") or norm_mime in (
        "application/x-ndjson",
        "application/jsonlines",
    ):
        try:
            text = content.decode("utf-8", errors="replace")
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            if not lines:
                return None, {}

            records: List[Dict[str, Any]] = []
            for line in lines:
                try:
                    parsed = json.loads(line)
                    if isinstance(parsed, dict):
                        records.append(parsed)
                except Exception:
                    pass

            if not records:
                return None, {}

            # Collect unique ordered keys
            headers_dict: Dict[str, bool] = {}
            for r in records[:50]:
                for k in r.keys():
                    headers_dict[str(k)] = True
            headers = list(headers_dict.keys()) or ["value"]

            total_rows = len(records)
            preview_rows = records[:25]

            md_lines = [
                f"**Dataset:** `{filename}`",
                f"**Format:** JSON Lines (JSONL) | **Total Rows:** {total_rows:,} | **Columns ({len(headers)}):** {', '.join(headers[:15])}",
                "",
                "| " + " | ".join(headers) + " |",
                "| " + " | ".join(["---"] * len(headers)) + " |",
            ]
            for r in preview_rows:
                cells = [
                    str(r.get(k, "")).replace("|", "\\|").replace("\n", " ").strip()
                    for k in headers
                ]
                md_lines.append("| " + " | ".join(cells) + " |")

            if total_rows > len(preview_rows):
                md_lines.append(
                    f"\n*(Showing top {len(preview_rows)} of {total_rows:,} total rows)*"
                )

            return "\n".join(md_lines), {
                "format": "jsonl",
                "row_count": total_rows,
                "column_count": len(headers),
                "columns": headers,
            }
        except Exception as e:
            logger.warning("Failed to parse JSONL file", filename=filename, error=str(e))
            return None, {"error": str(e)}

    # 3. CSV / TSV (.csv, .tsv, .tab)
    delimiter = "\t" if ext in (".tsv", ".tab") or norm_mime == "text/tab-separated-values" else ","
    try:
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1", errors="replace")

        # Sniff delimiter if extension is csv but may use semicolon
        sample = text[:2048]
        if ext == ".csv" and ";" in sample and "," not in sample:
            delimiter = ";"

        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        raw_rows = [r for r in reader if any(cell.strip() for cell in r)]
        if not raw_rows:
            return None, {}

        headers = [h.strip() or f"Column_{idx + 1}" for idx, h in enumerate(raw_rows[0])]
        data_rows = raw_rows[1:]
        total_rows = len(data_rows)
        preview_rows = data_rows[:25]

        fmt_name = "TSV" if delimiter == "\t" else "CSV"
        md_lines = [
            f"**Dataset:** `{filename}`",
            f"**Format:** {fmt_name} | **Total Rows:** {total_rows:,} | **Columns ({len(headers)}):** {', '.join(headers[:15])}",
            "",
            "| " + " | ".join(headers) + " |",
            "| " + " | ".join(["---"] * len(headers)) + " |",
        ]
        for r in preview_rows:
            padded = (r + [""] * len(headers))[: len(headers)]
            cleaned = [cell.replace("|", "\\|").replace("\n", " ").strip() for cell in padded]
            md_lines.append("| " + " | ".join(cleaned) + " |")

        if total_rows > len(preview_rows):
            md_lines.append(f"\n*(Showing top {len(preview_rows)} of {total_rows:,} total rows)*")

        return "\n".join(md_lines), {
            "format": fmt_name.lower(),
            "delimiter": delimiter,
            "row_count": total_rows,
            "column_count": len(headers),
            "columns": headers,
        }
    except Exception as e:
        logger.warning("Failed to parse CSV/TSV file", filename=filename, error=str(e))
        return None, {"error": str(e)}


class FileService:
    """
    Service managing workspace asset library files, local disk persistence,
    and metadata validation.
    """

    def __init__(self, base_storage_dir: Optional[Path] = None):
        self.storage_dir = base_storage_dir or BASE_STORAGE_DIR

    async def _ensure_workspace_dir(self, workspace_id: str) -> Path:
        ws_dir = self.storage_dir / "workspaces" / workspace_id / "files"
        p = AsyncPath(ws_dir)
        await p.mkdir(parents=True, exist_ok=True)
        return ws_dir

    async def save_file(
        self,
        db: AsyncSession,
        workspace_id: str,
        filename: str,
        content: bytes,
        mime_type: str,
    ) -> WorkspaceFile:
        """
        Validate, save file to disk, and persist record to workspace_files.
        """
        # 1. Verify workspace exists
        ws_result = await db.execute(select(Workspace).where(Workspace.id == workspace_id))
        ws = ws_result.scalar_one_or_none()
        if not ws:
            raise ValueError(f"Workspace '{workspace_id}' not found.")

        # 2. Validate payload size
        if len(content) > MAX_FILE_SIZE_BYTES:
            raise ValueError(f"File size ({len(content)} bytes) exceeds maximum limit of 20MB.")

        # 3. Categorize file & extract text
        normalized_mime = mime_type.lower().split(";")[0].strip()
        ext = os.path.splitext(filename.lower())[1]
        extracted_text: Optional[str] = None
        extra_meta: Dict[str, Any] = {}

        if ext == ".pdf" or normalized_mime == "application/pdf":
            category = "document"
            try:
                reader = pypdf.PdfReader(io.BytesIO(content))
                pages_text: List[str] = []
                for i, page in enumerate(reader.pages):
                    pt = page.extract_text() or ""
                    if pt.strip():
                        pages_text.append(f"--- Page {i + 1} ---\n{pt.strip()}")
                extracted_text = "\n\n".join(pages_text)
                extra_meta["page_count"] = len(reader.pages)
                extra_meta["is_encrypted"] = reader.is_encrypted
            except Exception as e:
                logger.warning("Failed to extract PDF text", filename=filename, error=str(e))
        elif ext in TABULAR_EXTENSIONS or normalized_mime in {
            "text/csv",
            "text/tab-separated-values",
            "application/x-ndjson",
            "application/jsonlines",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }:
            category = "tabular"
            text_summary, tab_meta = parse_tabular_bytes(filename, content, normalized_mime)
            extracted_text = text_summary
            extra_meta.update(tab_meta)
        elif ext in CODE_EXTENSIONS or normalized_mime in {
            "application/json",
            "application/javascript",
            "text/javascript",
            "application/x-yaml",
            "text/yaml",
            "text/x-python",
        }:
            category = "code"
            try:
                extracted_text = content.decode("utf-8", errors="replace")
            except Exception:
                pass
        elif ext in TEXT_DOC_EXTENSIONS or normalized_mime.startswith("text/"):
            category = "document"
            try:
                extracted_text = content.decode("utf-8", errors="replace")
            except Exception:
                pass
        elif normalized_mime in ALLOWED_IMAGE_MIMES or normalized_mime.startswith("image/"):
            category = "image"
        else:
            category = "other"
            try:
                decoded = content.decode("utf-8")
                if "\x00" not in decoded:
                    category = "document"
                    extracted_text = decoded
            except Exception:
                pass

        # 4. Write to workspace storage
        file_id = f"file_{uuid.uuid4().hex[:12]}"
        safe_name = _sanitize_filename(filename)
        dest_dir = await self._ensure_workspace_dir(workspace_id)
        storage_filename = f"{file_id}_{safe_name}"
        dest_path = dest_dir / storage_filename

        p = AsyncPath(dest_path)
        await p.write_bytes(content)

        # 5. Build database record
        file_record = WorkspaceFile(
            id=file_id,
            workspace_id=workspace_id,
            name=filename.strip() or safe_name,
            size_bytes=len(content),
            mime_type=normalized_mime,
            file_category=category,
            storage_path=str(dest_path),
            extracted_text=extracted_text,
            metadata_payload={
                "original_filename": filename,
                "sanitized_filename": safe_name,
                **extra_meta,
            },
        )

        db.add(file_record)
        await db.flush()

        logger.info(
            "Saved workspace file",
            file_id=file_id,
            workspace_id=workspace_id,
            size_bytes=len(content),
            mime_type=normalized_mime,
            category=category,
        )
        return file_record

    async def get_file(
        self,
        db: AsyncSession,
        workspace_id: str,
        file_id: str,
    ) -> Optional[WorkspaceFile]:
        """Fetch file metadata by workspace ID and file ID."""
        stmt = select(WorkspaceFile).where(
            WorkspaceFile.workspace_id == workspace_id,
            WorkspaceFile.id == file_id,
        )
        result = await db.execute(stmt)
        return result.scalar_one_or_none()

    async def list_files(
        self,
        db: AsyncSession,
        workspace_id: str,
        category: Optional[str] = None,
        limit: int = 50,
        offset: int = 0,
    ) -> List[WorkspaceFile]:
        """List files in workspace with optional category filter and pagination."""
        stmt = (
            select(WorkspaceFile)
            .where(WorkspaceFile.workspace_id == workspace_id)
            .order_by(WorkspaceFile.created_at.desc())
            .limit(limit)
            .offset(offset)
        )
        if category:
            stmt = stmt.where(WorkspaceFile.file_category == category)

        result = await db.execute(stmt)
        return list(result.scalars().all())

    async def delete_file(
        self,
        db: AsyncSession,
        workspace_id: str,
        file_id: str,
    ) -> bool:
        """Delete file from disk and database record."""
        file_record = await self.get_file(db, workspace_id, file_id)
        if not file_record:
            return False

        # Remove physical file from disk
        try:
            p = AsyncPath(file_record.storage_path)
            if await p.exists():
                await p.unlink()
        except Exception as e:
            logger.warning(
                "Could not delete physical file asset",
                path=file_record.storage_path,
                error=str(e),
            )

        await db.delete(file_record)
        await db.flush()
        logger.info("Deleted workspace file", file_id=file_id, workspace_id=workspace_id)
        return True


# Global service instance
file_service = FileService()
